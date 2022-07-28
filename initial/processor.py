from openpyxl import Workbook, load_workbook
import os

# os.path.exists("./sheets/expenses.xlsx")


class Ledger:
    def __init__(self, wb_path):
        is_exist = os.path.exists(wb_path)
        if not is_exist:
            print("ERROR: 404")
            return None

        self.wb = load_workbook(wb_path)
        self.data = {}
        for sheet in self.wb.sheetnames:
            self.data[sheet] = self.generate_expenses_dict(self.wb[sheet])

    def generate_expenses_dict(self, ws):
        expenses = {}

        for column in range(1, ws.max_column + 1):
            columnData = []
            for row in range(2, ws.max_row + 1):
                value = ws.cell(row, column).value
                if type(value) == int or type(value) == float:
                    columnData.append(value)
            expenses[ws.cell(1, column).value] = columnData
        return expenses

    def get(self):
        return self.data.keys()

    def _sum(self, array):
        sum = 0
        for value in array:
            sum += value
        return sum

    def generateStatement(self, save_location):
        expenseSum = {}
        for month in self.data:
            monthData = {}
            for expense in self.data[month]:
                totalExpense = self._sum(self.data[month][expense])
                if expense != None:
                    monthData[expense] = totalExpense
            expenseSum[month] = monthData

        # generate headers
        headers = []
        for month in expenseSum:
            for expense in list(expenseSum[month].keys()):
                if expense not in headers:
                    headers.append(expense)

        wb = Workbook()
        ws = wb.active
        ws.title = "STATEMENT OF INCOME"

        # append months
        months = [""]
        for month in expenseSum:
            months.append(month)
        ws.append(months)

        for header in headers:
            month_equivalent = []
            for month in expenseSum:
                month_equivalent.append(expenseSum[month].get(header, 0))
            ws.append([header] + month_equivalent)

        try:
            wb.save(save_location)
            return 200
        except PermissionError:
            return 401


# income = Ledger(
#     "C:\\Users\\Faruq\\Documents\\github\\manSys\\initial\\sheets\\expenses.xlsx")
